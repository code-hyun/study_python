import openpyxl
import crawling_test
from openpyxl.styles import Alignment, Font

def write_jobs_to_excel(jobs, filename="jobs.xlsx"):
    # Workbook 객체 생성
    excel_file = openpyxl.Workbook()
    excel_sheet = excel_file.active

    wrap_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    cell_font = Font(size=12)  # 나머지 셀의 폰트 크기 변경
    
    # 첫 번째 행에 열 이름 작성
    excel_sheet["A1"] = "회사 이름"
    excel_sheet["B1"] = "공고 이름"
    excel_sheet["C1"] = "채용 조건"
    excel_sheet["D1"] = "마감일"
    
    excel_sheet["A1"].font = header_font
    excel_sheet["B1"].font = header_font
    excel_sheet["C1"].font = header_font
    excel_sheet["D1"].font = header_font
    
    excel_sheet.column_dimensions['A'].width = 30
    excel_sheet.column_dimensions['B'].width = 80
    excel_sheet.column_dimensions['C'].width = 80
    excel_sheet.column_dimensions['D'].width = 50
    

    # 각 작업을 행으로 작성
    for index, job in enumerate(jobs, start=2):
        # 공고 내용을 파싱하여 각 항목을 분리합니다.
        job_split = job.split("\n")
        company_name = job_split[0].replace("회사이름 : ", "").strip()
        posting_name = job_split[1].replace("공고이름 : ", "").strip()
        job_condition = job_split[2].replace("채용조건 : ", "").strip()
        deadline = job_split[3].replace("마감일 : ", "").strip()

        # 워크시트에 행을 추가합니다.
        excel_sheet.cell(row=index, column=1, value=company_name)
        excel_sheet.cell(row=index, column=2, value=posting_name)
        excel_sheet.cell(row=index, column=3, value=job_condition)
        excel_sheet.cell(row=index, column=4, value=deadline)

        # 나머지 셀의 폰트 크기 변경
        for col in range(1, 5):
            excel_sheet.cell(row=index, column=col).font = cell_font
    
    # 변경 사항을 저장하고 엑셀 파일을 닫습니다.
    excel_file.save(filename)
    excel_file.close()
    
jobs = crawling_test.get_jobs()
write_jobs_to_excel(jobs, "jobs.xlsx")
